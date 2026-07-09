"""读取表格主图，直接执行 1688 以图搜图，并仅保存本地结果文件。"""

from __future__ import annotations

import argparse
import csv
import json
import multiprocessing as mp
from queue import Empty
import re
import sys
import time
import traceback
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import requests
from playwright.sync_api import sync_playwright


PROJECT_ROOT = Path(__file__).resolve().parent.parent
SRC_ROOT = PROJECT_ROOT / "src"

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))

from config.settings import get_settings
from ozon_selection.services.alibaba_image_search_pipeline import AlibabaImageSearchPipeline


DEFAULT_CSV_NAME = "Seerfar-Product20260706_2000.csv"
DEFAULT_IMAGE_COLUMN = "主图"
DEFAULT_TITLE_COLUMN = "标题"
DEFAULT_URL_COLUMN = "详情页地址"
DEFAULT_SKU_COLUMN = "SKU"
DEFAULT_SALES_COLUMN = "销量"
DEFAULT_PRICE_COLUMN = "售价"
DEFAULT_WEIGHT_COLUMN = "重量"
DEFAULT_SHOP_COLUMN = "店铺"
DEFAULT_BRAND_COLUMN = "品牌"
DEFAULT_CATEGORY_COLUMN = "类目"
DEFAULT_RESUME_STATE_NAME = "csv_1688_image_search_resume_state.json"
SUPPORTED_TABLE_SUFFIXES = {".csv", ".xlsx", ".xls"}
DEFAULT_CNY_TO_RUB_EXCHANGE_RATE = 11.0
IMAGE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36"
    ),
    "Referer": "https://www.ozon.ru/",
}


def parse_args() -> argparse.Namespace:
    """解析命令行参数。"""

    parser = argparse.ArgumentParser(description="读取 CSV 或 Excel 主图，直接执行 1688 以图搜图，并保存本地 xlsx/json。")
    parser.add_argument(
        "--csv",
        type=str,
        default=DEFAULT_CSV_NAME,
        help=f"CSV 文件路径，默认 `{DEFAULT_CSV_NAME}`。",
    )
    parser.add_argument(
        "--excel",
        type=str,
        default="",
        help="Excel 文件路径；传入后优先读取该文件。",
    )
    parser.add_argument(
        "--sheet-name",
        type=str,
        default="",
        help="可选：Excel 工作表名称；仅在读取 Excel 时生效。",
    )
    parser.add_argument(
        "--max-products",
        type=int,
        default=None,
        help="只处理前 N 个有效商品。",
    )
    parser.add_argument(
        "--max-results",
        type=int,
        default=None,
        help="每个商品最多抓前 N 个 1688 结果。",
    )
    parser.add_argument(
        "--download-dir",
        type=str,
        default="data/raw/csv_source_images",
        help="下载 CSV 主图到本地的根目录。",
    )
    parser.add_argument(
        "--background",
        action="store_true",
        help="后台模式运行 1688 浏览器。",
    )
    parser.add_argument(
        "--bitbrowser-browser-ids",
        type=str,
        default="",
        help="多个 BitBrowser 浏览器窗口 ID，逗号分隔；每个 worker 绑定一个独立 1688 账号。",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=None,
        help="并发 worker 数；默认等于传入的 BitBrowser 窗口数。",
    )
    parser.add_argument(
        "--no-resume",
        action="store_true",
        help="忽略上次进度，从头开始处理当前 CSV。",
    )
    return parser.parse_args()


def sanitize_filename(value: str, *, fallback: str) -> str:
    """把任意文本清洗为安全文件名。"""

    normalized = re.sub(r"\s+", "_", str(value or "").strip())
    normalized = re.sub(r"[^0-9A-Za-z_\-\u4e00-\u9fff]+", "", normalized)
    normalized = normalized.strip("_")
    return normalized or fallback


def parse_numeric(value: Any) -> int | None:
    """从字符串中提取整数。"""

    text = str(value or "").strip()
    if not text:
        return None
    matched = re.search(r"(\d+(?:[.,]\d+)?)", text.replace(" ", ""))
    if not matched:
        return None
    raw = matched.group(1).replace(",", ".")
    try:
        return int(float(raw))
    except ValueError:
        return None


def parse_float(value: Any) -> float | None:
    """从字符串中提取浮点数。"""

    text = str(value or "").strip()
    if not text:
        return None
    matched = re.search(r"(\d+(?:[.,]\d+)?)", text.replace(" ", ""))
    if not matched:
        return None
    raw = matched.group(1).replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return None


def parse_weight_to_grams(value: Any) -> int | None:
    """把原表重量解析为克。默认无单位时按克处理。"""

    text = str(value or "").strip().lower()
    if not text:
        return None

    normalized = text.replace(" ", "")
    numeric_value = parse_float(normalized)
    if numeric_value is None or numeric_value <= 0:
        return None

    if any(unit in normalized for unit in ("kg", "kgs", "公斤", "千克")):
        return int(round(numeric_value * 1000))
    if "斤" in normalized:
        return int(round(numeric_value * 500))
    return int(round(numeric_value))


def calculate_profit_margin(
    *,
    sale_price_rub: Any,
    weight_text: Any,
    unit_price_cny: Any,
    exchange_rate: float = DEFAULT_CNY_TO_RUB_EXCHANGE_RATE,
) -> tuple[float | None, str]:
    """按指定运费规则计算利润率。"""

    sale_price_text = str(sale_price_rub or "").strip()
    weight_text_value = str(weight_text or "").strip()
    unit_price_text = str(unit_price_cny or "").strip()
    sale_price_value = parse_float(sale_price_rub)
    weight_grams = parse_weight_to_grams(weight_text)
    unit_price_value = parse_float(unit_price_cny)
    if sale_price_value is None or sale_price_value <= 0:
        reason = "原表售价缺失" if not sale_price_text else f"原表售价无效: {sale_price_text}"
        return None, reason
    if weight_grams is None or weight_grams <= 0:
        reason = "原表重量缺失" if not weight_text_value else f"原表重量无效: {weight_text_value}"
        return None, reason
    if unit_price_value is None or unit_price_value < 0:
        reason = "1688单价缺失" if not unit_price_text else f"1688单价无效: {unit_price_text}"
        return None, reason

    shipping_fee_cny: float | None = None
    rule_label = ""
    if sale_price_value <= 1500 and 1 <= weight_grams <= 500:
        shipping_fee_cny = 3 + 0.035 * weight_grams
        rule_label = "规则1"
    elif 1501 <= sale_price_value <= 7000 and 1 <= weight_grams <= 2000:
        shipping_fee_cny = 16 + 0.035 * weight_grams
        rule_label = "规则2"

    if shipping_fee_cny is None:
        return None, (
            "售价与重量组合不适用当前运费规则: "
            f"售价={sale_price_value:.2f}卢布, 重量={weight_grams}g"
        )

    margin = (
        sale_price_value - (shipping_fee_cny + unit_price_value) * exchange_rate
    ) / sale_price_value
    return margin, rule_label


def resolve_source_path(*, csv_path: str, excel_path: str) -> Path:
    """解析输入表格路径。"""

    raw_path = excel_path.strip() or csv_path.strip()
    path = Path(raw_path).expanduser()
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    path = path.resolve()
    if not path.exists():
        raise FileNotFoundError(f"未找到表格文件: {path}")
    if path.suffix.lower() not in SUPPORTED_TABLE_SUFFIXES:
        raise ValueError(f"不支持的表格格式: {path.suffix}，仅支持 CSV/XLS/XLSX。")
    return path


def resolve_download_root(raw_path: str) -> Path:
    """解析图片下载根目录。"""

    path = Path(raw_path).expanduser()
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    return path.resolve()


def download_image(image_url: str, destination: Path) -> None:
    """下载主图到本地。"""

    if destination.exists():
        return

    destination.parent.mkdir(parents=True, exist_ok=True)
    response = requests.get(
        image_url,
        headers=IMAGE_HEADERS,
        timeout=20,
        allow_redirects=True,
    )
    response.raise_for_status()
    destination.write_bytes(response.content)


def get_resume_state_path(settings) -> Path:
    """返回 CSV 图搜图续跑状态文件路径。"""

    return settings.processed_data_path / DEFAULT_RESUME_STATE_NAME


def load_resume_state(state_path: Path) -> dict[str, Any]:
    """读取续跑状态文件。"""

    if not state_path.exists():
        return {}
    try:
        return json.loads(state_path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_resume_state(state_path: Path, payload: dict[str, Any]) -> None:
    """写入续跑状态文件。"""

    state_path.parent.mkdir(parents=True, exist_ok=True)
    state_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def build_resume_key(source_path: Path) -> str:
    """生成续跑状态中的表格唯一键。"""

    return str(source_path.resolve())


def get_resume_checkpoint(state_path: Path, source_path: Path) -> dict[str, Any]:
    """获取指定表格文件的续跑断点。"""

    state = load_resume_state(state_path)
    checkpoint = state.get(build_resume_key(source_path))
    return checkpoint if isinstance(checkpoint, dict) else {}


def reset_resume_checkpoint(state_path: Path, *, source_path: Path) -> None:
    """清空指定表格文件的续跑断点。"""

    state = load_resume_state(state_path)
    state.pop(build_resume_key(source_path), None)
    save_resume_state(state_path, state)


def read_table_rows(source_path: Path, *, sheet_name: str = "") -> list[dict[str, Any]]:
    """读取 CSV 或 Excel 行。"""

    suffix = source_path.suffix.lower()
    if suffix == ".csv":
        with source_path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))

    frame = pd.read_excel(
        source_path,
        sheet_name=sheet_name or 0,
    )
    frame = frame.where(pd.notna(frame), "")
    return frame.to_dict(orient="records")


def build_table_products(
    source_path: Path,
    download_root: Path,
    *,
    max_products: int | None,
    skip_valid_products: int = 0,
    skip_completed_valid_products: set[int] | None = None,
    sheet_name: str = "",
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """从表格构造 1688 图搜图所需的 Ozon 商品结构。"""

    products: list[dict[str, Any]] = []
    seen_skus: set[str] = set()
    stats = {
        "total_rows": 0,
        "valid_rows": 0,
        "scanned_valid_products": 0,
        "skipped_resumed_products": 0,
        "downloaded_images": 0,
        "skipped_missing_image": 0,
        "skipped_missing_sales": 0,
        "skipped_duplicate_sku": 0,
        "skipped_completed_products": 0,
        "download_failed": 0,
    }
    completed_valid_products = skip_completed_valid_products or set()

    batch_keyword = sanitize_filename(source_path.stem, fallback="table_source")
    target_dir = download_root / batch_keyword
    rows = read_table_rows(source_path, sheet_name=sheet_name)
    for row_index, row in enumerate(rows, start=1):
        stats["total_rows"] += 1

        image_url = str(row.get(DEFAULT_IMAGE_COLUMN) or "").strip()
        if not image_url:
            stats["skipped_missing_image"] += 1
            continue

        monthly_sales = parse_numeric(row.get(DEFAULT_SALES_COLUMN))
        if monthly_sales is None or monthly_sales <= 0:
            stats["skipped_missing_sales"] += 1
            continue

        raw_sku = str(row.get(DEFAULT_SKU_COLUMN) or "").strip()
        sku = sanitize_filename(raw_sku, fallback=f"table_{row_index}")
        if sku in seen_skus:
            stats["skipped_duplicate_sku"] += 1
            continue

        seen_skus.add(sku)
        stats["scanned_valid_products"] += 1
        if stats["scanned_valid_products"] <= skip_valid_products:
            stats["skipped_resumed_products"] += 1
            continue
        if stats["scanned_valid_products"] in completed_valid_products:
            stats["skipped_completed_products"] += 1
            continue

        image_path = target_dir / sku / "1.jpg"
        try:
            before_exists = image_path.exists()
            download_image(image_url, image_path)
            if not before_exists and image_path.exists():
                stats["downloaded_images"] += 1
        except Exception as exc:
            stats["download_failed"] += 1
            print(f"[csv-1688] image download failed sku={sku} url={image_url} error={exc}", flush=True)
            continue

        stats["valid_rows"] += 1
        products.append(
            {
                "validProductIndex": stats["scanned_valid_products"],
                "sku": sku,
                "name": str(row.get(DEFAULT_TITLE_COLUMN) or "").strip(),
                "url": str(row.get(DEFAULT_URL_COLUMN) or "").strip(),
                "imageUrl": image_url,
                "localImagePath": str(image_path),
                "price": parse_numeric(row.get(DEFAULT_PRICE_COLUMN)),
                "csvPriceText": str(row.get(DEFAULT_PRICE_COLUMN) or "").strip(),
                "csvWeightText": str(row.get(DEFAULT_WEIGHT_COLUMN) or "").strip(),
                "monthlySales": monthly_sales,
                "dailySales": None,
                "batchKeyword": batch_keyword,
                "brand": str(row.get(DEFAULT_BRAND_COLUMN) or "").strip(),
                "category": str(row.get(DEFAULT_CATEGORY_COLUMN) or "").strip(),
                "shop": str(row.get(DEFAULT_SHOP_COLUMN) or "").strip(),
                "attributes": [],
                "rawPayload": row,
            }
        )
        if max_products is not None and max_products > 0 and len(products) >= max_products:
            break

    return products, stats


def parse_bitbrowser_browser_ids(raw_value: str) -> list[str]:
    """解析 BitBrowser 浏览器窗口 ID 列表。"""

    parts = re.split(r"[\s,]+", raw_value.strip())
    browser_ids: list[str] = []
    seen: set[str] = set()
    for part in parts:
        normalized = part.strip()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        browser_ids.append(normalized)
    return browser_ids


def build_worker_browser_ids(
    *,
    cli_value: str,
    settings,
    requested_workers: int | None,
) -> list[str]:
    """确定每个 worker 对应的 BitBrowser 窗口 ID。"""

    browser_ids = parse_bitbrowser_browser_ids(cli_value)
    if not browser_ids and settings.alibaba1688_bitbrowser_browser_id.strip():
        browser_ids = [settings.alibaba1688_bitbrowser_browser_id.strip()]

    if requested_workers is not None and requested_workers <= 0:
        raise ValueError("--workers 必须大于 0。")
    if requested_workers is not None:
        if requested_workers == 1:
            return browser_ids[:1]
        if not browser_ids:
            raise ValueError("启用多进程前请通过 --bitbrowser-browser-ids 提供 BitBrowser 窗口 ID。")
        if requested_workers > len(browser_ids):
            raise ValueError("--workers 不能大于 BitBrowser 窗口 ID 数量。")
        browser_ids = browser_ids[:requested_workers]

    return browser_ids


def normalize_completed_valid_products(raw_value: Any) -> list[int]:
    """把续跑状态里的离散完成索引转成整数列表。"""

    if not isinstance(raw_value, list):
        return []

    normalized: set[int] = set()
    for value in raw_value:
        try:
            index = int(value)
        except (TypeError, ValueError):
            continue
        if index > 0:
            normalized.add(index)
    return sorted(normalized)


def update_resume_checkpoint(
    state_path: Path,
    *,
    source_path: Path,
    completed_valid_products: int | list[int] | tuple[int, ...] | set[int] | None,
    last_completed_sku: str,
    last_run_id: str,
) -> dict[str, Any]:
    """更新指定表格文件的续跑断点，支持乱序完成索引。"""

    state = load_resume_state(state_path)
    key = build_resume_key(source_path)
    existing_checkpoint = state.get(key) if isinstance(state.get(key), dict) else {}
    contiguous_completed = int(existing_checkpoint.get("processed_valid_products") or 0)
    pending_completed = set(
        normalize_completed_valid_products(existing_checkpoint.get("completed_valid_products"))
    )

    new_completed: list[int]
    if completed_valid_products is None:
        new_completed = []
    elif isinstance(completed_valid_products, int):
        new_completed = [completed_valid_products]
    else:
        new_completed = []
        for value in completed_valid_products:
            try:
                normalized = int(value)
            except (TypeError, ValueError):
                continue
            if normalized > 0:
                new_completed.append(normalized)

    for index in new_completed:
        if index <= contiguous_completed:
            continue
        pending_completed.add(index)

    while contiguous_completed + 1 in pending_completed:
        contiguous_completed += 1
        pending_completed.remove(contiguous_completed)

    state[key] = {
        "source_reference": str(source_path),
        "source_mtime_ns": source_path.stat().st_mtime_ns,
        "processed_valid_products": contiguous_completed,
        "completed_valid_products": sorted(pending_completed),
        "last_completed_sku": last_completed_sku,
        "last_run_id": last_run_id,
        "updated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    save_resume_state(state_path, state)
    return state[key]


def build_worker_settings_update(base_settings, *, worker_index: int, browser_id: str | None) -> dict[str, Any]:
    """为每个 worker 生成独立的 1688 运行配置。"""

    worker_tag = sanitize_filename(browser_id or f"worker_{worker_index + 1}", fallback=f"worker_{worker_index + 1}")
    return {
        "alibaba1688_headless": base_settings.alibaba1688_headless,
        "alibaba1688_bitbrowser_browser_id": browser_id or base_settings.alibaba1688_bitbrowser_browser_id,
        "alibaba1688_auth_state_file": f"auth-state-1688-{worker_tag}.json",
        "alibaba1688_user_data_dir": f"browser-profile-1688-{worker_tag}",
    }


def split_products_for_workers(products: list[dict[str, Any]], worker_count: int) -> list[list[dict[str, Any]]]:
    """按轮询方式分配商品，尽量均衡每个 worker 的负载。"""

    buckets: list[list[dict[str, Any]]] = [[] for _ in range(worker_count)]
    for index, product in enumerate(products):
        buckets[index % worker_count].append(product)
    return [bucket for bucket in buckets if bucket]


def build_worker_counters() -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    """初始化单个 worker 的聚合计数器。"""

    image_prefilter_result = {
        "status": "skipped",
        "evaluated_items": 0,
        "passed_items": 0,
        "failed_items": 0,
        "selected_items": 0,
    }
    search_result = {
        "status": "completed",
        "failed_products": 0,
    }
    detail_result = {
        "status": "skipped",
        "enriched_items": 0,
        "failed_items": 0,
    }
    return image_prefilter_result, search_result, detail_result


def merge_status(current_status: str, next_status: str) -> str:
    """合并多个 worker 的阶段状态。"""

    status_order = {
        "failed": 4,
        "partial_failed": 3,
        "completed": 2,
        "verified": 2,
        "saved": 2,
        "skipped": 1,
        "pending": 0,
    }
    current_score = status_order.get(current_status, 0)
    next_score = status_order.get(next_status, 0)
    return next_status if next_score >= current_score else current_status


def print_completed_progress(
    *,
    completed_count: int,
    total_count: int,
    global_index: int,
    sku: str,
    worker_label: str | None = None,
) -> None:
    """打印当前已完成的商品进度。"""

    prefix = f"[csv-1688][{worker_label}]" if worker_label else "[csv-1688]"
    print(
        f"{prefix} completed: {completed_count}/{total_count} "
        f"latest_global_valid_index={global_index} SKU={sku}",
        flush=True,
    )


def run_product_batch(
    *,
    settings,
    products: list[dict[str, Any]],
    max_results: int | None,
    progress_queue: mp.queues.Queue | None = None,
    worker_label: str = "worker-1",
) -> dict[str, Any]:
    """在单个浏览器会话中顺序处理分配到的商品。"""

    pipeline = AlibabaImageSearchPipeline(settings=settings)
    results: list[dict[str, Any]] = []
    completed_results: list[dict[str, Any]] = []
    attempted_valid_indices: list[int] = []
    image_prefilter_result, search_result, detail_result = build_worker_counters()
    auth_state_path = settings.alibaba1688_auth_state_path

    with sync_playwright() as playwright:
        session, auth_state_path = pipeline.browser_search.ensure_ready_context(playwright)
        context = session.context
        working_page = context.new_page()
        try:
            total_products = len(products)
            print(f"[csv-1688][{worker_label}] assigned products: {total_products}", flush=True)
            for index, product in enumerate(products, start=1):
                global_index = int(product.get("validProductIndex") or index)
                attempted_valid_indices.append(global_index)
                print(
                    f"[csv-1688][{worker_label}] progress: {index}/{total_products} "
                    f"global_valid_index={global_index} SKU={product.get('sku')}",
                    flush=True,
                )
                try:
                    image_result = pipeline.browser_search.search_by_uploaded_image_in_context(
                        context,
                        image_path=product["localImagePath"],
                        page=working_page,
                        max_results=max_results,
                        enrich_details=False,
                    )
                except Exception as exc:
                    search_result["status"] = "partial_failed"
                    search_result["failed_products"] += 1
                    print(
                        f"[csv-1688][{worker_label}] search failed for SKU={product.get('sku')}: {exc}",
                        flush=True,
                    )
                    results.append(
                        pipeline.build_failed_search_result(product=product, error=str(exc))
                    )
                    if progress_queue is not None:
                        progress_queue.put(
                            {
                                "type": "progress",
                                "worker_label": worker_label,
                                "valid_index": global_index,
                                "sku": str(product.get("sku") or ""),
                                "worker_completed_count": len(results),
                                "worker_total_count": total_products,
                            }
                        )
                    continue

                next_working_page = image_result.pop("_active_page", None)
                if next_working_page is not None and next_working_page is not working_page:
                    try:
                        if not working_page.is_closed():
                            working_page.close()
                    except Exception:
                        pass
                    working_page = next_working_page

                prefilter = pipeline.compare_result_images_with_ai(product, image_result["items"])
                image_prefilter_result["status"] = merge_status(
                    image_prefilter_result["status"],
                    prefilter["status"],
                )
                image_prefilter_result["evaluated_items"] += prefilter.get("evaluated_items", 0)
                image_prefilter_result["passed_items"] += prefilter.get("passed_items", 0)
                image_prefilter_result["failed_items"] += prefilter.get("failed_items", 0)

                best_items = pipeline.select_best_image_match_items(image_result["items"])
                image_prefilter_result["selected_items"] += len(best_items)
                if best_items:
                    try:
                        pipeline.browser_search.enrich_results_with_detail(context, best_items)
                        detail_result["status"] = merge_status(detail_result["status"], "completed")
                        detail_result["enriched_items"] += len(best_items)
                    except Exception as exc:
                        detail_result["status"] = merge_status(detail_result["status"], "partial_failed")
                        detail_result["failed_items"] += len(best_items)
                        for item in best_items:
                            item["detail_error"] = str(exc)

                image_result["items"] = best_items
                result_entry = {
                    "ozon_product": product,
                    "image_search": image_result,
                }
                results.append(result_entry)
                if best_items:
                    completed_results.append(result_entry)

                if progress_queue is not None:
                    progress_queue.put(
                        {
                            "type": "progress",
                            "worker_label": worker_label,
                            "valid_index": global_index,
                            "sku": str(product.get("sku") or ""),
                            "worker_completed_count": len(results),
                            "worker_total_count": total_products,
                        }
                    )
        finally:
            try:
                working_page.close()
            except Exception:
                pass
            session.close()

    return {
        "worker_label": worker_label,
        "browser_id": settings.alibaba1688_bitbrowser_browser_id,
        "auth_state_path": str(auth_state_path),
        "results": results,
        "completed_results": completed_results,
        "attempted_valid_indices": attempted_valid_indices,
        "image_prefilter_result": image_prefilter_result,
        "search_result": search_result,
        "detail_result": detail_result,
    }


def worker_main(
    *,
    worker_index: int,
    browser_id: str | None,
    products: list[dict[str, Any]],
    max_results: int | None,
    base_settings_update: dict[str, Any],
    progress_queue: mp.queues.Queue,
) -> None:
    """子进程入口。"""

    worker_label = f"worker-{worker_index + 1}"
    try:
        settings = get_settings().model_copy(deep=True, update=base_settings_update)
        payload = run_product_batch(
            settings=settings,
            products=products,
            max_results=max_results,
            progress_queue=progress_queue,
            worker_label=worker_label,
        )
        progress_queue.put(
            {
                "type": "done",
                "worker_index": worker_index,
                "worker_label": worker_label,
                "browser_id": browser_id or "",
                "payload": payload,
            }
        )
    except Exception as exc:
        progress_queue.put(
            {
                "type": "worker_error",
                "worker_index": worker_index,
                "worker_label": worker_label,
                "browser_id": browser_id or "",
                "error": str(exc),
                "traceback": traceback.format_exc(),
            }
        )


def run_products_in_parallel(
    *,
    settings,
    products: list[dict[str, Any]],
    browser_ids: list[str],
    max_results: int | None,
    resume_state_path: Path,
    source_path: Path,
    run_id: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any], dict[str, Any], dict[str, Any], list[dict[str, Any]], list[str]]:
    """使用多进程 + 多 BitBrowser 窗口执行 1688 图搜图。"""

    assignments = split_products_for_workers(products, len(browser_ids))
    ctx = mp.get_context("spawn")
    progress_queue = ctx.Queue()
    processes: list[mp.context.SpawnProcess] = []
    worker_payloads: list[dict[str, Any]] = []
    worker_errors: list[dict[str, Any]] = []
    terminal_workers: set[int] = set()
    completed_valid_indices: set[int] = set()
    total_products = len(products)

    for worker_index, assigned_products in enumerate(assignments):
        browser_id = browser_ids[worker_index]
        settings_update = build_worker_settings_update(
            settings,
            worker_index=worker_index,
            browser_id=browser_id,
        )
        process = ctx.Process(
            target=worker_main,
            kwargs={
                "worker_index": worker_index,
                "browser_id": browser_id,
                "products": assigned_products,
                "max_results": max_results,
                "base_settings_update": settings_update,
                "progress_queue": progress_queue,
            },
        )
        process.start()
        processes.append(process)
        print(
            f"[csv-1688] started worker-{worker_index + 1} browser_id={browser_id} products={len(assigned_products)}",
            flush=True,
        )

    while len(terminal_workers) < len(processes):
        try:
            message = progress_queue.get(timeout=1)
        except Empty:
            if not any(process.is_alive() for process in processes):
                break
            continue

        message_type = message.get("type")
        if message_type == "progress":
            valid_index = int(message.get("valid_index") or 0)
            update_resume_checkpoint(
                resume_state_path,
                source_path=source_path,
                completed_valid_products=valid_index,
                last_completed_sku=str(message.get("sku") or ""),
                last_run_id=run_id,
            )
            if valid_index > 0 and valid_index not in completed_valid_indices:
                completed_valid_indices.add(valid_index)
                print_completed_progress(
                    completed_count=len(completed_valid_indices),
                    total_count=total_products,
                    global_index=valid_index,
                    sku=str(message.get("sku") or ""),
                    worker_label=str(message.get("worker_label") or ""),
                )
            continue

        worker_index = int(message.get("worker_index") or 0)
        terminal_workers.add(worker_index)
        if message_type == "done":
            worker_payloads.append(message["payload"])
        elif message_type == "worker_error":
            worker_errors.append(
                {
                    "worker_index": worker_index,
                    "worker_label": message.get("worker_label"),
                    "browser_id": message.get("browser_id"),
                    "error": message.get("error"),
                    "traceback": message.get("traceback"),
                }
            )

    for process in processes:
        process.join()

    while True:
        try:
            message = progress_queue.get_nowait()
        except Empty:
            break

        message_type = message.get("type")
        if message_type == "progress":
            valid_index = int(message.get("valid_index") or 0)
            update_resume_checkpoint(
                resume_state_path,
                source_path=source_path,
                completed_valid_products=valid_index,
                last_completed_sku=str(message.get("sku") or ""),
                last_run_id=run_id,
            )
            if valid_index > 0 and valid_index not in completed_valid_indices:
                completed_valid_indices.add(valid_index)
                print_completed_progress(
                    completed_count=len(completed_valid_indices),
                    total_count=total_products,
                    global_index=valid_index,
                    sku=str(message.get("sku") or ""),
                    worker_label=str(message.get("worker_label") or ""),
                )
            continue

        worker_index = int(message.get("worker_index") or 0)
        terminal_workers.add(worker_index)
        if message_type == "done":
            worker_payloads.append(message["payload"])
        elif message_type == "worker_error":
            worker_errors.append(
                {
                    "worker_index": worker_index,
                    "worker_label": message.get("worker_label"),
                    "browser_id": message.get("browser_id"),
                    "error": message.get("error"),
                    "traceback": message.get("traceback"),
                }
            )

    for worker_index, process in enumerate(processes):
        if worker_index in terminal_workers:
            continue
        if process.exitcode not in (0, None):
            worker_errors.append(
                {
                    "worker_index": worker_index,
                    "worker_label": f"worker-{worker_index + 1}",
                    "browser_id": browser_ids[worker_index],
                    "error": f"worker exited with code {process.exitcode}",
                    "traceback": "",
                }
            )

    results: list[dict[str, Any]] = []
    completed_results: list[dict[str, Any]] = []
    image_prefilter_result, search_result, detail_result = build_worker_counters()
    auth_state_paths: list[str] = []
    worker_summaries: list[dict[str, Any]] = []

    for payload in worker_payloads:
        results.extend(payload["results"])
        completed_results.extend(payload["completed_results"])
        auth_state_path = str(payload.get("auth_state_path") or "")
        if auth_state_path:
            auth_state_paths.append(auth_state_path)
        image_prefilter_result["status"] = merge_status(
            image_prefilter_result["status"],
            str(payload["image_prefilter_result"].get("status") or "skipped"),
        )
        image_prefilter_result["evaluated_items"] += int(payload["image_prefilter_result"].get("evaluated_items") or 0)
        image_prefilter_result["passed_items"] += int(payload["image_prefilter_result"].get("passed_items") or 0)
        image_prefilter_result["failed_items"] += int(payload["image_prefilter_result"].get("failed_items") or 0)
        image_prefilter_result["selected_items"] += int(payload["image_prefilter_result"].get("selected_items") or 0)
        search_result["status"] = merge_status(
            search_result["status"],
            str(payload["search_result"].get("status") or "completed"),
        )
        search_result["failed_products"] += int(payload["search_result"].get("failed_products") or 0)
        detail_result["status"] = merge_status(
            detail_result["status"],
            str(payload["detail_result"].get("status") or "skipped"),
        )
        detail_result["enriched_items"] += int(payload["detail_result"].get("enriched_items") or 0)
        detail_result["failed_items"] += int(payload["detail_result"].get("failed_items") or 0)
        worker_summaries.append(
            {
                "worker_label": payload.get("worker_label"),
                "browser_id": payload.get("browser_id"),
                "auth_state_path": payload.get("auth_state_path"),
                "processed_attempts": len(payload["results"]),
                "processed_products": len(payload["completed_results"]),
                "attempted_valid_indices": payload.get("attempted_valid_indices") or [],
                "search_failed_products": payload["search_result"].get("failed_products"),
            }
        )
        update_resume_checkpoint(
            resume_state_path,
            source_path=source_path,
            completed_valid_products=payload.get("attempted_valid_indices"),
            last_completed_sku="",
            last_run_id=run_id,
        )

    if worker_errors:
        search_result["status"] = merge_status(search_result["status"], "partial_failed")
        detail_result["status"] = merge_status(detail_result["status"], "partial_failed")

    results.sort(key=lambda item: int(item["ozon_product"].get("validProductIndex") or 0))
    completed_results.sort(key=lambda item: int(item["ozon_product"].get("validProductIndex") or 0))
    worker_summaries.sort(key=lambda item: str(item.get("worker_label") or ""))

    if auth_state_paths:
        deduped_paths = list(dict.fromkeys(auth_state_paths))
    else:
        deduped_paths = []

    return (
        results,
        completed_results,
        image_prefilter_result,
        search_result,
        detail_result,
        worker_summaries + worker_errors,
        deduped_paths,
    )


def build_csv_excel_rows(
    pipeline: AlibabaImageSearchPipeline,
    results: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """构造 CSV 主图搜图专用 Excel 行，保留原表售价和重量。"""

    rows: list[dict[str, Any]] = []
    for result in results:
        ozon_product = result["ozon_product"]
        items = result["image_search"]["items"]
        for index, item in enumerate(items, start=1):
            image_comparison = item.get("ai_image_comparison") or {}
            profit_margin, profit_margin_note = calculate_profit_margin(
                sale_price_rub=ozon_product.get("csvPriceText") or ozon_product.get("price"),
                weight_text=ozon_product.get("csvWeightText"),
                unit_price_cny=item.get("unit_price"),
            )
            profit_margin_value: float | str = profit_margin if profit_margin is not None else profit_margin_note
            rows.append(
                {
                    "Ozon SKU": ozon_product.get("sku"),
                    "Ozon商品": ozon_product.get("name"),
                    "原表售价": ozon_product.get("csvPriceText"),
                    "原表重量": ozon_product.get("csvWeightText"),
                    "利润率": profit_margin_value,
                    "Ozon月销量": ozon_product.get("monthlySales"),
                    "Ozon日销量": ozon_product.get("dailySales"),
                    "Ozon属性数": len(ozon_product.get("attributes") or []),
                    "Ozon商品属性": pipeline.format_attributes(ozon_product.get("attributes")),
                    "Ozon链接": ozon_product.get("url"),
                    "Ozon主图路径": ozon_product.get("localImagePath"),
                    "1688序号": index,
                    "1688标题": item.get("title"),
                    "1688链接": item.get("detail_url"),
                    "1688价格": item.get("price"),
                    "1688价格文案": item.get("price_text"),
                    "1688单价": item.get("unit_price"),
                    "1688单价文案": item.get("unit_price_text"),
                    "1688重量文案": item.get("weight_text"),
                    "1688重量(g)": item.get("weight_grams"),
                    "1688属性数": len(item.get("attributes") or []),
                    "1688商品属性": pipeline.format_attributes(item.get("attributes")),
                    "1688卖家": item.get("seller"),
                    "GPT主图状态": image_comparison.get("status"),
                    "GPT主图是否同款": pipeline.format_bool_value(image_comparison.get("same_product")),
                    "GPT主图同款分": image_comparison.get("image_match_score"),
                    "GPT主图置信度": image_comparison.get("confidence"),
                    "GPT主图说明": image_comparison.get("summary"),
                    "详情抓取异常": item.get("detail_error"),
                }
            )
    return rows


def export_csv_excel(settings, rows: list[dict[str, Any]]) -> dict[str, Any]:
    """导出 CSV 主图搜图 Excel。"""

    if not rows:
        return {
            "status": "skipped",
            "count": 0,
            "reason": "empty_rows",
        }

    output_dir = settings.ozon_scrape_output_path
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"alibaba1688_image_search_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    frame = pd.DataFrame(rows)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="1688图搜图", index=False)

    workbook = load_workbook(output_path)
    sheet = workbook["1688图搜图"]
    widths = [14, 30, 12, 14, 12, 12, 12, 10, 42, 40, 42, 10, 42, 42, 12, 16, 12, 16, 16, 14, 10, 42, 18, 12, 12, 10, 40, 12, 24]
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[AlibabaImageSearchPipeline.column_letter(column_index)].width = width
    for row_index in range(2, sheet.max_row + 1):
        cell = sheet.cell(row_index, 5)
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.00%"

    if sheet.max_row > 2:
        group_start = 2
        previous_key = tuple(sheet.cell(2, column_index).value for column_index in range(1, 12))
        for row_index in range(3, sheet.max_row + 1):
            current_key = tuple(sheet.cell(row_index, column_index).value for column_index in range(1, 12))
            if current_key != previous_key:
                merge_csv_excel_group(sheet, group_start, row_index - 1)
                group_start = row_index
                previous_key = current_key
        merge_csv_excel_group(sheet, group_start, sheet.max_row)

    workbook.save(output_path)
    return {
        "status": "saved",
        "count": len(rows),
        "path": str(output_path),
    }


def merge_csv_excel_group(sheet, start_row: int, end_row: int) -> None:
    """按 Ozon 商品分组合并 CSV 专用导出中的 Ozon 列。"""

    if end_row <= start_row:
        return

    for column_index in range(1, 12):
        sheet.merge_cells(
            start_row=start_row,
            start_column=column_index,
            end_row=end_row,
            end_column=column_index,
        )
        cell = sheet.cell(start_row, column_index)
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def build_source_type(source_path: Path) -> str:
    """根据输入文件后缀生成 source_type。"""

    return f"{source_path.suffix.lower().removeprefix('.')} _main_images".replace(" ", "")


def save_report_json(
    *,
    output_dir: Path,
    run_id: str,
    source_path: Path,
    auth_state_paths: list[str],
    excel_path: str,
    stats: dict[str, Any],
    results: list[dict[str, Any]],
    worker_summaries: list[dict[str, Any]] | None = None,
) -> Path:
    """保存 JSON 报告。"""

    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / f"alibaba1688_image_search_{run_id}.json"
    primary_auth_state_path = auth_state_paths[0] if auth_state_paths else ""
    payload = {
        "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "source_type": build_source_type(source_path),
        "source_reference": str(source_path),
        "auth_state_path": primary_auth_state_path,
        "auth_state_paths": auth_state_paths,
        "excel_path": excel_path,
        "stats": stats,
        "results": results,
    }
    if worker_summaries:
        payload["worker_summaries"] = worker_summaries
    report_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return report_path


def main() -> None:
    """读取表格主图，执行 1688 图搜图。"""

    args = parse_args()
    source_path = resolve_source_path(csv_path=args.csv, excel_path=args.excel)
    download_root = resolve_download_root(args.download_dir)
    settings = get_settings().model_copy(
        deep=True,
        update={"alibaba1688_headless": bool(args.background)},
    )
    resume_state_path = get_resume_state_path(settings)
    if args.no_resume:
        reset_resume_checkpoint(resume_state_path, source_path=source_path)
    resume_checkpoint = {} if args.no_resume else get_resume_checkpoint(resume_state_path, source_path)
    checkpoint_mtime_ns = int(resume_checkpoint.get("source_mtime_ns") or 0)
    if resume_checkpoint and checkpoint_mtime_ns and checkpoint_mtime_ns != source_path.stat().st_mtime_ns:
        print("[csv-1688] source file changed, ignoring old resume checkpoint.", flush=True)
        reset_resume_checkpoint(resume_state_path, source_path=source_path)
        resume_checkpoint = {}
    resume_offset = int(resume_checkpoint.get("processed_valid_products") or 0)
    completed_resume_indices = set(
        normalize_completed_valid_products(resume_checkpoint.get("completed_valid_products"))
    )
    if args.no_resume:
        resume_offset = 0
        completed_resume_indices.clear()

    run_id = time.strftime("%Y%m%d_%H%M%S")
    products, load_stats = build_table_products(
        source_path,
        download_root,
        max_products=args.max_products,
        skip_valid_products=resume_offset,
        skip_completed_valid_products=completed_resume_indices,
        sheet_name=args.sheet_name,
    )
    if not products:
        if (
            resume_offset > 0
            or completed_resume_indices
            or load_stats["scanned_valid_products"] <= resume_offset
        ):
            print(f"[csv-1688] source file: {source_path}", flush=True)
            print(f"[csv-1688] resume checkpoint: already processed {resume_offset} valid products", flush=True)
            print("[csv-1688] no remaining valid products to process.", flush=True)
            return
        raise RuntimeError("表格中没有可用于 1688 图搜图的有效商品。")

    browser_ids = build_worker_browser_ids(
        cli_value=args.bitbrowser_browser_ids,
        settings=settings,
        requested_workers=args.workers,
    )
    if browser_ids and len(browser_ids) > len(products):
        browser_ids = browser_ids[:len(products)]
    worker_count = len(browser_ids) if len(browser_ids) > 1 else 1

    results: list[dict[str, Any]] = []
    completed_results: list[dict[str, Any]] = []
    image_prefilter_result, search_result, detail_result = build_worker_counters()
    auth_state_paths: list[str] = []
    worker_summaries: list[dict[str, Any]] = []

    print(f"[csv-1688] source file: {source_path}", flush=True)
    print(f"[csv-1688] queued valid products: {len(products)}", flush=True)
    if args.sheet_name.strip():
        print(f"[csv-1688] excel sheet: {args.sheet_name.strip()}", flush=True)
    if resume_offset > 0:
        print(
            f"[csv-1688] resume from valid product #{resume_offset + 1} "
            f"(already processed: {resume_offset})",
            flush=True,
        )
    if completed_resume_indices:
        print(
            f"[csv-1688] resume completed sparse indices: {len(completed_resume_indices)}",
            flush=True,
        )

    if worker_count > 1:
        print(
            f"[csv-1688] parallel mode enabled: workers={worker_count} browser_ids={','.join(browser_ids)}",
            flush=True,
        )
        (
            results,
            completed_results,
            image_prefilter_result,
            search_result,
            detail_result,
            worker_summaries,
            auth_state_paths,
        ) = run_products_in_parallel(
            settings=settings,
            products=products,
            browser_ids=browser_ids,
            max_results=args.max_results,
            resume_state_path=resume_state_path,
            source_path=source_path,
            run_id=run_id,
        )
    else:
        if browser_ids:
            settings = settings.model_copy(
                deep=True,
                update=build_worker_settings_update(
                    settings,
                    worker_index=0,
                    browser_id=browser_ids[0],
                ),
            )
        pipeline = AlibabaImageSearchPipeline(settings=settings)
        with sync_playwright() as playwright:
            session, auth_state_path = pipeline.browser_search.ensure_ready_context(playwright)
            auth_state_paths = [str(auth_state_path)]
            context = session.context
            working_page = context.new_page()
            try:
                total_products = len(products)
                for index, product in enumerate(products, start=1):
                    global_index = int(product.get("validProductIndex") or (resume_offset + index))
                    print(
                        f"[csv-1688] progress: {index}/{total_products} global_valid_index={global_index} "
                        f"SKU={product.get('sku')} name={product.get('name') or 'unknown'}",
                        flush=True,
                    )
                    try:
                        image_result = pipeline.browser_search.search_by_uploaded_image_in_context(
                            context,
                            image_path=product["localImagePath"],
                            page=working_page,
                            max_results=args.max_results,
                            enrich_details=False,
                        )
                    except Exception as exc:
                        search_result["status"] = "partial_failed"
                        search_result["failed_products"] += 1
                        print(
                            f"[csv-1688] search failed for SKU={product.get('sku')}: {exc}",
                            flush=True,
                        )
                        results.append(
                            pipeline.build_failed_search_result(product=product, error=str(exc))
                        )
                        update_resume_checkpoint(
                            resume_state_path,
                            source_path=source_path,
                            completed_valid_products=global_index,
                            last_completed_sku=str(product.get("sku") or ""),
                            last_run_id=run_id,
                        )
                        print_completed_progress(
                            completed_count=len(results),
                            total_count=total_products,
                            global_index=global_index,
                            sku=str(product.get("sku") or ""),
                        )
                        continue

                    next_working_page = image_result.pop("_active_page", None)
                    if next_working_page is not None and next_working_page is not working_page:
                        try:
                            if not working_page.is_closed():
                                working_page.close()
                        except Exception:
                            pass
                        working_page = next_working_page

                    prefilter = pipeline.compare_result_images_with_ai(product, image_result["items"])
                    image_prefilter_result["status"] = merge_status(
                        image_prefilter_result["status"],
                        prefilter["status"],
                    )
                    image_prefilter_result["evaluated_items"] += prefilter.get("evaluated_items", 0)
                    image_prefilter_result["passed_items"] += prefilter.get("passed_items", 0)
                    image_prefilter_result["failed_items"] += prefilter.get("failed_items", 0)

                    best_items = pipeline.select_best_image_match_items(image_result["items"])
                    image_prefilter_result["selected_items"] += len(best_items)
                    if best_items:
                        try:
                            pipeline.browser_search.enrich_results_with_detail(context, best_items)
                            detail_result["status"] = merge_status(detail_result["status"], "completed")
                            detail_result["enriched_items"] += len(best_items)
                        except Exception as exc:
                            detail_result["status"] = merge_status(detail_result["status"], "partial_failed")
                            detail_result["failed_items"] += len(best_items)
                            for item in best_items:
                                item["detail_error"] = str(exc)

                    image_result["items"] = best_items
                    result_entry = {
                        "ozon_product": product,
                        "image_search": image_result,
                    }
                    results.append(result_entry)
                    if best_items:
                        completed_results.append(result_entry)
                    update_resume_checkpoint(
                        resume_state_path,
                        source_path=source_path,
                        completed_valid_products=global_index,
                        last_completed_sku=str(product.get("sku") or ""),
                        last_run_id=run_id,
                    )
                    print_completed_progress(
                        completed_count=len(results),
                        total_count=total_products,
                        global_index=global_index,
                        sku=str(product.get("sku") or ""),
                    )
            finally:
                try:
                    working_page.close()
                except Exception:
                    pass
                session.close()

        worker_summaries = [
            {
                "worker_label": "worker-1",
                "browser_id": settings.alibaba1688_bitbrowser_browser_id,
                "auth_state_path": auth_state_paths[0] if auth_state_paths else "",
                "processed_attempts": len(results),
                "processed_products": len(completed_results),
                "attempted_valid_indices": [
                    int(item["ozon_product"].get("validProductIndex") or 0)
                    for item in results
                ],
                "search_failed_products": search_result["failed_products"],
            }
        ]

    pipeline = AlibabaImageSearchPipeline(settings=settings)
    excel_rows = build_csv_excel_rows(pipeline, completed_results)
    excel_result = export_csv_excel(settings, excel_rows)
    report_stats = {
        **load_stats,
        "resume_offset": resume_offset,
        "resume_completed_sparse_count": len(completed_resume_indices),
        "worker_count": worker_count,
        "worker_error_count": sum(1 for item in worker_summaries if item.get("error")),
        "processed_products": len(completed_results),
        "processed_attempts": len(results),
        "matched_items": sum(len(result["image_search"]["items"]) for result in completed_results),
        "search_failed_products": search_result["failed_products"],
        "image_prefilter_result": image_prefilter_result,
        "detail_result": detail_result,
    }
    report_path = save_report_json(
        output_dir=settings.ozon_scrape_output_path,
        run_id=run_id,
        source_path=source_path,
        auth_state_paths=auth_state_paths,
        excel_path=str(excel_result.get("path") or ""),
        stats=report_stats,
        results=results,
        worker_summaries=worker_summaries,
    )

    print("1688 csv image search: completed")
    print(f"source_type: {build_source_type(source_path)}")
    print(f"source_reference: {source_path}")
    print(f"resume_offset: {resume_offset}")
    print(f"valid_products: {load_stats['valid_rows']}")
    print(f"worker_count: {worker_count}")
    print(f"processed_attempts: {len(results)}")
    print(f"processed_products: {len(completed_results)}")
    print(f"matched_items: {report_stats['matched_items']}")
    print(f"downloaded_images: {load_stats['downloaded_images']}")
    print(f"search_status: {search_result['status']}")
    print(f"search_failed_products: {search_result['failed_products']}")
    print(f"image_prefilter_status: {image_prefilter_result['status']}")
    print(f"image_prefilter_items: {image_prefilter_result['evaluated_items']}")
    print(f"image_prefilter_passed: {image_prefilter_result['passed_items']}")
    print(f"image_prefilter_selected: {image_prefilter_result['selected_items']}")
    print(f"detail_status: {detail_result['status']}")
    print(f"detail_enriched_items: {detail_result['enriched_items']}")
    print(f"excel_status: {excel_result['status']}")
    if excel_result.get("path"):
        print(f"excel_path: {excel_result['path']}")
    if excel_result.get("reason"):
        print(f"excel_note: {excel_result['reason']}")
    if excel_result.get("error"):
        print(f"excel_error: {excel_result['error']}")
    print(f"json_path: {report_path}")


if __name__ == "__main__":
    main()
