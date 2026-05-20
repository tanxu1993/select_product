"""读取 Ozon 选品 Excel，对通过商品按主图做 GPT 去重，并导出新 Excel。"""

from __future__ import annotations

import argparse
import hashlib
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import httpx
from openpyxl import Workbook, load_workbook
from PIL import Image


PROJECT_ROOT = Path(__file__).resolve().parent.parent
SRC_ROOT = PROJECT_ROOT / "src"

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))

from config.settings import get_settings
from ozon_selection.services.product_comparison import ProductComparisonService


PASS_LABEL = "✅ 通过"


@dataclass(slots=True)
class PassedProductRow:
    """单条通过商品记录。"""

    source_row_index: int
    data: dict[str, Any]
    image_path: Path
    image_md5: str
    image_ahash: int
    normalized_name: str

    @property
    def sku(self) -> str:
        return str(self.data.get("SKU") or "")

    @property
    def name(self) -> str:
        return str(self.data.get("商品名称") or "")


@dataclass(slots=True)
class DedupDecision:
    """去重判断结果。"""

    group_id: int
    kept: bool
    representative_sku: str
    reason: str
    image_match_score: int | None = None
    confidence: str | None = None
    summary: str | None = None
    same_product: bool | None = None


@dataclass(slots=True)
class ProductGroup:
    """去重分组。"""

    group_id: int
    representative: PassedProductRow
    members: list[PassedProductRow] = field(default_factory=list)
    representative_reason: str = "initial_representative"


class OzonPassedProductsDeduper:
    """对通过商品按主图做去重。"""

    MAX_COMPARE_RETRIES = 4
    RETRY_SLEEP_SECONDS = 5

    def __init__(self) -> None:
        self.settings = get_settings()
        self.comparison_service = ProductComparisonService(settings=self.settings)
        if not self.comparison_service.is_configured:
            raise RuntimeError("OPENAI_API_KEY 未配置，无法执行 GPT 主图去重。")

    def run(self, input_path: Path, *, output_path: Path | None = None) -> Path:
        """执行去重并输出 Excel。"""

        headers, passed_rows = self.load_passed_rows(input_path)
        if not passed_rows:
            raise RuntimeError("输入 Excel 中没有 `✅ 通过` 的商品。")

        print(f"[dedupe] input: {input_path}", flush=True)
        print(f"[dedupe] passed rows: {len(passed_rows)}", flush=True)

        groups: list[ProductGroup] = []
        decisions: dict[int, DedupDecision] = {}
        total = len(passed_rows)

        for index, row in enumerate(passed_rows, start=1):
            print(f"[dedupe] progress: {index}/{total} SKU={row.sku} name={row.name}", flush=True)
            group, decision = self.assign_row_to_group(row, groups)
            if decision.kept and group.representative.source_row_index == row.source_row_index:
                group.members.append(row)
            elif not decision.kept:
                group.members.append(row)
            decisions[row.source_row_index] = decision

        output_file = output_path or self.build_output_path(input_path)
        self.write_output_workbook(
            output_path=output_file,
            headers=headers,
            groups=groups,
            decisions=decisions,
        )
        print(f"[dedupe] groups kept: {len(groups)}", flush=True)
        print(f"[dedupe] output: {output_file}", flush=True)
        return output_file

    def assign_row_to_group(
        self,
        row: PassedProductRow,
        groups: list[ProductGroup],
    ) -> tuple[ProductGroup, DedupDecision]:
        """把商品分配到现有分组，或创建新分组。"""

        if not groups:
            group = ProductGroup(group_id=1, representative=row)
            groups.append(group)
            return group, DedupDecision(
                group_id=group.group_id,
                kept=True,
                representative_sku=row.sku,
                reason="first_item",
            )

        candidates = self.find_candidate_groups(row, groups)
        for group in candidates:
            comparison = self.compare_rows_with_gpt(row, group.representative)
            if self.is_same_product(comparison):
                return group, DedupDecision(
                    group_id=group.group_id,
                    kept=False,
                    representative_sku=group.representative.sku,
                    reason="gpt_duplicate",
                    image_match_score=comparison.get("image_match_score"),
                    confidence=comparison.get("confidence"),
                    summary=comparison.get("summary"),
                    same_product=comparison.get("same_product"),
                )

        group = ProductGroup(group_id=len(groups) + 1, representative=row)
        groups.append(group)
        return group, DedupDecision(
            group_id=group.group_id,
            kept=True,
            representative_sku=row.sku,
            reason="new_unique_group",
        )

    def find_candidate_groups(self, row: PassedProductRow, groups: list[ProductGroup]) -> list[ProductGroup]:
        """基于本地图片和标题相似度筛选 GPT 候选分组。"""

        exact_hash_matches: list[ProductGroup] = []
        fuzzy_matches: list[tuple[int, ProductGroup]] = []

        for group in groups:
            representative = group.representative
            if row.image_md5 == representative.image_md5:
                exact_hash_matches.append(group)
                continue

            distance = self.hamming_distance(row.image_ahash, representative.image_ahash)
            same_name = row.normalized_name and row.normalized_name == representative.normalized_name
            similar_name = (
                row.normalized_name
                and representative.normalized_name
                and (
                    row.normalized_name in representative.normalized_name
                    or representative.normalized_name in row.normalized_name
                )
            )
            if distance <= 6 or same_name or (distance <= 10 and similar_name):
                fuzzy_matches.append((distance, group))

        if exact_hash_matches:
            return exact_hash_matches

        fuzzy_matches.sort(key=lambda item: item[0])
        return [group for _, group in fuzzy_matches[:8]]

    def compare_rows_with_gpt(
        self,
        row: PassedProductRow,
        representative: PassedProductRow,
    ) -> dict[str, Any]:
        """使用 GPT 对两张 Ozon 主图做同款判断。"""

        for attempt in range(1, self.MAX_COMPARE_RETRIES + 1):
            try:
                result = self.comparison_service.compare_product_images(
                    ozon_product={
                        "name": row.name,
                        "localImagePath": str(row.image_path),
                    },
                    supplier_product={
                        "title": representative.name,
                        "image_url": str(representative.image_path),
                    },
                )
                print(
                    "[dedupe] compare "
                    f"{row.sku} -> {representative.sku} "
                    f"status={result.get('status')} same={result.get('same_product')} "
                    f"score={result.get('image_match_score')}",
                    flush=True,
                )
                return result
            except (httpx.HTTPError, httpx.TimeoutException) as exc:
                if attempt >= self.MAX_COMPARE_RETRIES:
                    print(
                        "[dedupe] compare failed permanently "
                        f"{row.sku} -> {representative.sku}: {exc}",
                        flush=True,
                    )
                    return {
                        "status": "failed",
                        "error": str(exc),
                    }
                sleep_seconds = self.RETRY_SLEEP_SECONDS * attempt
                print(
                    "[dedupe] compare retry "
                    f"{attempt}/{self.MAX_COMPARE_RETRIES} for {row.sku} -> {representative.sku} "
                    f"after error: {exc}",
                    flush=True,
                )
                time.sleep(sleep_seconds)

    @staticmethod
    def is_same_product(comparison: dict[str, Any]) -> bool:
        """判断 GPT 是否认定为重复商品。"""

        if comparison.get("status") != "completed":
            return False
        if comparison.get("same_product") is True:
            return True
        return int(comparison.get("image_match_score") or 0) >= 95

    def load_passed_rows(self, input_path: Path) -> tuple[list[str], list[PassedProductRow]]:
        """读取 Excel 中通过的商品行。"""

        workbook = load_workbook(input_path, read_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(value) for value in rows[0]]
        rows_index = {header: idx for idx, header in enumerate(headers)}

        passed_rows: list[PassedProductRow] = []
        for source_row_index, row_values in enumerate(rows[1:], start=2):
            if row_values[rows_index["结果"]] != PASS_LABEL:
                continue

            image_path = Path(str(row_values[rows_index["主图路径"]]))
            if not image_path.exists():
                print(f"[dedupe] skip row {source_row_index}, missing image: {image_path}", flush=True)
                continue

            row_data = {header: row_values[idx] for idx, header in enumerate(headers)}
            passed_rows.append(
                PassedProductRow(
                    source_row_index=source_row_index,
                    data=row_data,
                    image_path=image_path,
                    image_md5=self.compute_md5(image_path),
                    image_ahash=self.compute_average_hash(image_path),
                    normalized_name=self.normalize_name(str(row_data.get("商品名称") or "")),
                )
            )

        return headers, passed_rows

    def write_output_workbook(
        self,
        *,
        output_path: Path,
        headers: list[str],
        groups: list[ProductGroup],
        decisions: dict[int, DedupDecision],
    ) -> None:
        """写出去重后的工作簿。"""

        workbook = Workbook()
        deduped_sheet = workbook.active
        deduped_sheet.title = "去重后结果"
        detail_sheet = workbook.create_sheet("去重明细")

        deduped_headers = headers + ["去重组ID", "去重组大小", "去重保留SKU", "去重保留原因"]
        detail_headers = headers + [
            "去重组ID",
            "是否保留",
            "保留SKU",
            "去重原因",
            "GPT主图去重分",
            "GPT置信度",
            "GPT主图说明",
        ]
        deduped_sheet.append(deduped_headers)
        detail_sheet.append(detail_headers)

        for group in groups:
            representative = group.representative
            group_size = len(group.members)
            decision = decisions[representative.source_row_index]
            deduped_sheet.append(
                [representative.data.get(header) for header in headers]
                + [group.group_id, group_size, representative.sku, decision.reason]
            )

            for member in group.members:
                member_decision = decisions[member.source_row_index]
                detail_sheet.append(
                    [member.data.get(header) for header in headers]
                    + [
                        member_decision.group_id,
                        "是" if member_decision.kept else "否",
                        member_decision.representative_sku,
                        member_decision.reason,
                        member_decision.image_match_score,
                        member_decision.confidence,
                        member_decision.summary,
                    ]
                )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

    def build_output_path(self, input_path: Path) -> Path:
        """构造输出文件路径。"""

        timestamp = time.strftime("%Y%m%d_%H%M%S")
        return input_path.parent / f"{input_path.stem}_gpt去重_{timestamp}.xlsx"

    @staticmethod
    def normalize_name(name: str) -> str:
        """归一化标题，去掉明显占位词。"""

        normalized = name.strip().lower()
        placeholders = [
            "новинка",
            "цена что надо",
            "осталась 1 шт",
        ]
        for placeholder in placeholders:
            normalized = normalized.replace(placeholder, " ")
        normalized = " ".join(normalized.split())
        return normalized

    @staticmethod
    def compute_md5(image_path: Path) -> str:
        """计算图片 MD5。"""

        return hashlib.md5(image_path.read_bytes()).hexdigest()

    @staticmethod
    def compute_average_hash(image_path: Path, *, size: int = 16) -> int:
        """计算简单感知哈希。"""

        image = Image.open(image_path).convert("L").resize((size, size))
        pixels = list(image.getdata())
        average = sum(pixels) / len(pixels)
        bits = "".join("1" if pixel >= average else "0" for pixel in pixels)
        return int(bits, 2)

    @staticmethod
    def hamming_distance(left: int, right: int) -> int:
        """计算感知哈希汉明距离。"""

        return (left ^ right).bit_count()


def parse_args() -> argparse.Namespace:
    """解析命令行参数。"""

    parser = argparse.ArgumentParser(description="对通过的 Ozon 商品按主图做 GPT 去重。")
    parser.add_argument(
        "input_excel",
        help="待去重的 Ozon 选品 Excel 路径。",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="可选输出路径；默认写到同目录新文件。",
    )
    return parser.parse_args()


def main() -> None:
    """命令行入口。"""

    args = parse_args()
    input_path = Path(args.input_excel).resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"未找到输入 Excel: {input_path}")

    output_path = Path(args.output).resolve() if args.output else None
    result = OzonPassedProductsDeduper().run(input_path, output_path=output_path)
    print(result)


if __name__ == "__main__":
    main()
