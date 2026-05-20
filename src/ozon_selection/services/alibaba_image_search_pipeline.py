"""1688 主图搜图流水线。"""

from __future__ import annotations

import json
import time
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from playwright.sync_api import sync_playwright

from config.settings import Settings, get_settings
from ozon_selection.collectors.alibaba.image_search import Alibaba1688ImageSearchBrowser
from ozon_selection.repositories.alibaba_image_search_repository import AlibabaImageSearchRepository
from ozon_selection.repositories.ozon_batch_repository import OzonBatchRepository
from ozon_selection.repositories.supplier_link_repository import SupplierLinkRepository
from ozon_selection.services.ozon_candidate_pipeline import OzonCandidatePipeline
from ozon_selection.services.product_comparison import ProductComparisonService


class AlibabaImageSearchPipeline:
    """负责读取 Ozon 主图并批量执行 1688 图搜图。"""

    EXCEL_OZON_MERGE_COLUMNS = (1, 2, 3, 4, 5, 6, 7, 8)
    OZON_PROMO_TITLE_MARKERS = (
        "новинка",
        "баллов за отзыв",
        "цена что надо",
        "вау-цены",
        "вау цены",
        "осталась 1 шт",
    )

    def __init__(
        self,
        settings: Settings | None = None,
        *,
        browser_search: Alibaba1688ImageSearchBrowser | None = None,
        repository: SupplierLinkRepository | None = None,
        batch_repository: OzonBatchRepository | None = None,
        sqlite_repository: AlibabaImageSearchRepository | None = None,
        comparison_service: ProductComparisonService | None = None,
    ) -> None:
        self.settings = settings or get_settings()
        self.browser_search = browser_search or Alibaba1688ImageSearchBrowser(settings=self.settings)
        self.repository = repository or SupplierLinkRepository(settings=self.settings)
        self.batch_repository = batch_repository or OzonBatchRepository(settings=self.settings)
        self.sqlite_repository = sqlite_repository or AlibabaImageSearchRepository(settings=self.settings)
        self.comparison_service = comparison_service or ProductComparisonService(settings=self.settings)

    def run(
        self,
        manifest_path: str | Path | None = None,
        *,
        max_products: int | None = None,
        max_results: int | None = None,
    ) -> dict[str, Any]:
        """执行本地主图上传图搜图流程。"""

        results: list[dict[str, Any]] = []
        completed_results: list[dict[str, Any]] = []
        image_prefilter_result = {
            "status": "skipped",
            "evaluated_items": 0,
            "passed_items": 0,
            "failed_items": 0,
            "selected_items": 0,
        }
        search_result = {
            "status": "completed",
            "failed_products": 0,
        }
        login_result = {
            "status": "pending",
        }
        detail_result = {
            "status": "skipped",
            "enriched_items": 0,
            "failed_items": 0,
        }
        processed_product_ids: list[int] = []
        with sync_playwright() as playwright:
            print("[1688] running login preflight...", flush=True)
            session, auth_state_path = self.browser_search.ensure_ready_context(playwright)
            context = session.context
            login_result = {
                "status": "verified",
                "auth_state_path": str(auth_state_path),
            }
            working_page = context.new_page()
            try:
                print("[1688] loading reviewed Ozon products...", flush=True)
                source = self.load_source_products(manifest_path)
                products = self.filter_products_with_images(source.get("products") or [])
                products = self.limit_products(products, max_products)
                if not products:
                    raise RuntimeError("Ozon 清单里没有可用的本地主图，无法执行 1688 图搜图。")
                print(f"[1688] source loaded: {source['source_reference']}", flush=True)
                print(f"[1688] products with local images: {len(products)}", flush=True)
                if max_results is not None:
                    print(f"[1688] per-product result limit: {max_results}", flush=True)

                total_products = len(products)
                for index, product in enumerate(products, start=1):
                    print(
                        f"[1688] progress: {index}/{total_products} "
                        f"SKU={product.get('sku')} name={product.get('name') or 'unknown'}",
                        flush=True,
                    )
                    print(
                        f"[1688] searching image for SKU={product.get('sku')} path={product.get('localImagePath')}",
                        flush=True,
                    )
                    try:
                        image_result = self.browser_search.search_by_uploaded_image_in_context(
                            context,
                            image_path=product["localImagePath"],
                            page=working_page,
                            max_results=max_results,
                            enrich_details=False,
                        )
                    except Exception as exc:
                        search_result["status"] = "partial_failed"
                        search_result["failed_products"] += 1
                        print(
                            f"[1688] search failed for SKU={product.get('sku')}: {exc}",
                            flush=True,
                        )
                        results.append(self.build_failed_search_result(product=product, error=str(exc)))
                        continue
                    next_working_page = image_result.pop("_active_page", None)
                    if next_working_page is not None and next_working_page is not working_page:
                        try:
                            if not working_page.is_closed():
                                working_page.close()
                        except Exception:
                            pass
                        working_page = next_working_page
                    print(
                        f"[1688] got {len(image_result['items'])} matches for SKU={product.get('sku')}",
                        flush=True,
                    )
                    prefilter = self.compare_result_images_with_ai(product, image_result["items"])
                    image_prefilter_result["status"] = prefilter["status"]
                    image_prefilter_result["evaluated_items"] += prefilter.get("evaluated_items", 0)
                    image_prefilter_result["passed_items"] += prefilter.get("passed_items", 0)
                    image_prefilter_result["failed_items"] += prefilter.get("failed_items", 0)
                    best_items = self.select_best_image_match_items(image_result["items"])
                    image_prefilter_result["selected_items"] += len(best_items)
                    if not best_items:
                        print(
                            f"[1688] no exportable match after image prefilter for SKU={product.get('sku')}",
                            flush=True,
                        )

                    if best_items:
                        print(
                            f"[1688] top image match candidates: {len(best_items)}",
                            flush=True,
                        )
                        try:
                            self.browser_search.enrich_results_with_detail(context, best_items)
                            detail_result["status"] = "completed"
                            detail_result["enriched_items"] += len(best_items)
                        except Exception as exc:
                            detail_result["status"] = "partial_failed"
                            detail_result["failed_items"] += len(best_items)
                            for item in best_items:
                                item["detail_error"] = str(exc)
                    image_result["items"] = best_items
                    result_entry = {
                        "ozon_product": product,
                        "image_search": image_result,
                    }
                    results.append(result_entry)
                    reviewed_product_id = int(product.get("reviewedProductId") or 0)
                    if best_items:
                        completed_results.append(result_entry)
                    if reviewed_product_id > 0 and best_items:
                        processed_product_ids.append(reviewed_product_id)
            finally:
                try:
                    working_page.close()
                except Exception:
                    pass
                session.close()

        db_payloads = self.build_database_rows(completed_results)
        excel_rows = self.build_excel_rows(completed_results)
        excel_result = self.save_to_excel(excel_rows)
        processed_result = self.mark_processed_products(
            processed_product_ids,
            source_type=source["source_type"],
        )
        sqlite_result = self.save_to_sqlite(db_payloads, source_reference=source["source_reference"])
        database_result = self.save_to_database(db_payloads)

        return {
            "source_type": source["source_type"],
            "source_reference": source["source_reference"],
            "processed_products": len(completed_results),
            "matched_items": sum(len(result["image_search"]["items"]) for result in completed_results),
            "auth_state_path": str(auth_state_path),
            "login_result": login_result,
            "excel_result": excel_result,
            "processed_result": processed_result,
            "sqlite_result": sqlite_result,
            "database_result": database_result,
            "search_result": search_result,
            "image_prefilter_result": image_prefilter_result,
            "detail_result": detail_result,
        }

    @staticmethod
    def build_failed_search_result(*, product: dict[str, Any], error: str) -> dict[str, Any]:
        """构造单个商品图搜失败后的占位结果，避免中断整批流程。"""

        return {
            "ozon_product": product,
            "image_search": {
                "search_url": "",
                "final_url": "",
                "logged_in": True,
                "image_path": str(product.get("localImagePath") or ""),
                "items": [],
                "search_error": error,
            },
        }

    def load_source_products(self, manifest_path: str | Path | None) -> dict[str, Any]:
        """优先读取人工审核完成后的 SQLite 商品；必要时兼容旧 manifest。"""

        if manifest_path:
            manifest = self.load_manifest(manifest_path)
            return {
                "source_type": "manifest",
                "source_reference": manifest["manifest_path"],
                "products": manifest.get("products") or [],
            }

        reviewed_products = self.load_reviewed_products_from_sqlite()
        if not reviewed_products:
            raise RuntimeError("SQLite 中没有已人工审核完成的 Ozon 商品，无法执行 1688 图搜图。")
        return {
            "source_type": "sqlite_reviewed",
            "source_reference": f"sqlite://{self.settings.sqlite_db_path}",
            "products": reviewed_products,
        }

    def load_manifest(self, manifest_path: str | Path | None) -> dict[str, Any]:
        """兼容读取第 2 步生成的 Ozon 候选商品清单。"""

        output_dir = self.settings.ozon_scrape_output_path
        path = Path(manifest_path) if manifest_path else OzonCandidatePipeline.find_latest_manifest(output_dir)
        payload = json.loads(path.read_text(encoding="utf-8"))
        payload["manifest_path"] = str(path)
        return payload

    def load_reviewed_products_from_sqlite(self) -> list[dict[str, Any]]:
        """读取 SQLite 中人工审核完成后保留的 Ozon 商品。"""

        rows = self.batch_repository.list_completed_products()
        products = [self.map_reviewed_product_to_ozon_payload(row) for row in rows]
        return [product for product in products if self.is_valid_ozon_source_product(product)]

    @staticmethod
    def map_reviewed_product_to_ozon_payload(row: dict[str, Any]) -> dict[str, Any]:
        """把 SQLite 审核后商品映射成 Ozon manifest 兼容结构。"""

        payload = dict(row.get("raw_payload") or {})
        payload.setdefault("sku", row.get("source_product_id"))
        payload.setdefault("reviewedProductId", row.get("id"))
        payload.setdefault("name", row.get("title"))
        payload.setdefault("detailTitle", row.get("detail_title"))
        payload.setdefault("url", row.get("source_url"))
        payload.setdefault("imageUrl", row.get("image_url"))
        payload.setdefault("localImagePath", row.get("image_path"))
        payload.setdefault("detailImageUrl", row.get("detail_image_url"))
        payload.setdefault("attributes", row.get("attributes") or [])
        payload.setdefault("price", row.get("price"))
        payload.setdefault("detailPrice", row.get("detail_price"))
        payload.setdefault("monthlySales", row.get("monthly_sales"))
        payload.setdefault("dailySales", row.get("daily_sales"))
        payload.setdefault("batchKeyword", row.get("batch_keyword"))
        payload.setdefault("batchId", row.get("batch_id"))
        payload.setdefault("batchManifestPath", row.get("batch_manifest_path"))
        return payload

    @staticmethod
    def filter_products_with_images(products: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """过滤掉没有本地图片的商品。"""

        filtered: list[dict[str, Any]] = []
        for product in products:
            image_path = product.get("localImagePath")
            if image_path and Path(image_path).exists():
                filtered.append(product)
        return filtered

    @classmethod
    def is_valid_ozon_source_product(cls, product: dict[str, Any]) -> bool:
        """过滤掉明显不合格的 Ozon 源商品，避免进入 1688 图搜图。"""

        monthly_sales = product.get("monthlySales")
        try:
            normalized_monthly_sales = int(monthly_sales) if monthly_sales is not None else None
        except (TypeError, ValueError):
            normalized_monthly_sales = None

        if normalized_monthly_sales is None or normalized_monthly_sales <= 0:
            return False

        title = str(product.get("name") or "").strip().lower()
        if not title:
            return False
        if any(marker in title for marker in cls.OZON_PROMO_TITLE_MARKERS):
            return False

        source_url = str(product.get("url") or "").strip()
        local_image_path = str(product.get("localImagePath") or "").strip()
        if not source_url or not local_image_path:
            return False

        return True

    @staticmethod
    def limit_products(products: list[dict[str, Any]], max_products: int | None) -> list[dict[str, Any]]:
        """按需截断待处理的 Ozon 商品数量。"""

        if max_products is None or max_products <= 0:
            return products
        return products[:max_products]

    @staticmethod
    def build_database_rows(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """构造 `supplier_links` 表写入记录。"""

        rows: list[dict[str, Any]] = []
        for result in results:
            ozon_product = result["ozon_product"]
            for item in result["image_search"]["items"]:
                image_comparison = item.get("ai_image_comparison") or {}
                rows.append(
                    {
                        "source_platform": "ozon",
                        "ozon_batch_id": ozon_product.get("batchId"),
                        "ozon_keyword": ozon_product.get("batchKeyword"),
                        "source_product_id": str(ozon_product.get("sku") or ""),
                        "source_title": ozon_product.get("name"),
                        "source_product_url": ozon_product.get("url"),
                        "source_image_url": ozon_product.get("imageUrl"),
                        "source_image_path": ozon_product.get("localImagePath"),
                        "source_price": ozon_product.get("price"),
                        "supplier_platform": "1688",
                        "supplier_title": item.get("title"),
                        "supplier_product_url": item.get("detail_url"),
                        "supplier_image_url": item.get("image_url"),
                        "supplier_price": item.get("price"),
                        "supplier_price_text": item.get("price_text"),
                        "supplier_unit_price": item.get("unit_price"),
                        "supplier_unit_price_text": item.get("unit_price_text"),
                        "supplier_weight_text": item.get("weight_text"),
                        "supplier_weight_grams": item.get("weight_grams"),
                        "supplier_attributes": item.get("attributes") or [],
                        "supplier_seller": item.get("seller"),
                        "ai_image_same_product": image_comparison.get("same_product"),
                        "ai_image_match_score": image_comparison.get("image_match_score"),
                        "ai_image_confidence": image_comparison.get("confidence"),
                        "ai_image_summary": image_comparison.get("summary"),
                        "search_method": "1688_uploaded_local_image",
                        "raw_payload": item,
                    }
                )
        return rows

    @staticmethod
    def build_excel_rows(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """构造 Excel 导出行。"""

        rows: list[dict[str, Any]] = []
        for result in results:
            ozon_product = result["ozon_product"]
            items = result["image_search"]["items"]
            for index, item in enumerate(items, start=1):
                image_comparison = item.get("ai_image_comparison") or {}
                rows.append(
                    {
                        "Ozon SKU": ozon_product.get("sku"),
                        "Ozon商品": ozon_product.get("name"),
                        "Ozon月销量": ozon_product.get("monthlySales"),
                        "Ozon日销量": ozon_product.get("dailySales"),
                        "Ozon属性数": len(ozon_product.get("attributes") or []),
                        "Ozon商品属性": AlibabaImageSearchPipeline.format_attributes(ozon_product.get("attributes")),
                        "Ozon链接": ozon_product.get("url"),
                        "Ozon主图路径": ozon_product.get("localImagePath"),
                        "1688序号": index,
                        "1688标题": item.get("title"),
                        "1688链接": item.get("detail_url"),
                        "1688价格": item.get("price"),
                        "1688价格文案": item.get("price_text"),
                        "1688单价": item.get("unit_price"),
                        "1688单价文案": item.get("unit_price_text"),
                        "1688重量文案": item.get("weight_text"),
                        "1688重量(g)": item.get("weight_grams"),
                        "1688属性数": len(item.get("attributes") or []),
                        "1688商品属性": AlibabaImageSearchPipeline.format_attributes(item.get("attributes")),
                        "1688卖家": item.get("seller"),
                        "GPT主图状态": image_comparison.get("status"),
                        "GPT主图是否同款": AlibabaImageSearchPipeline.format_bool_value(
                            image_comparison.get("same_product")
                        ),
                        "GPT主图同款分": image_comparison.get("image_match_score"),
                        "GPT主图置信度": image_comparison.get("confidence"),
                        "GPT主图说明": image_comparison.get("summary"),
                        "详情抓取异常": item.get("detail_error"),
                    }
                )
        return rows

    def export_to_excel(self, rows: list[dict[str, Any]]) -> Path:
        """导出 1688 图搜图结果 Excel。"""

        output_dir = self.settings.ozon_scrape_output_path
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"alibaba1688_image_search_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        frame = pd.DataFrame(rows)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            frame.to_excel(writer, sheet_name="1688图搜图", index=False)

        workbook = load_workbook(output_path)
        sheet = workbook["1688图搜图"]
        widths = [14, 30, 12, 12, 10, 42, 40, 42, 10, 42, 42, 12, 16, 12, 16, 16, 14, 10, 42, 18, 12, 12, 10, 40, 12, 12, 12, 10, 10, 40, 42, 24]
        for column_index, width in enumerate(widths, start=1):
            sheet.column_dimensions[self.column_letter(column_index)].width = width
        self.merge_ozon_columns(sheet)
        workbook.save(output_path)
        return output_path

    def merge_ozon_columns(self, sheet) -> None:
        """按 Ozon 商品分组合并重复的 Ozon 列。"""

        if sheet.max_row <= 2:
            return

        group_start = 2
        previous_key = self.build_ozon_merge_key(sheet, row=2)

        for row_index in range(3, sheet.max_row + 1):
            current_key = self.build_ozon_merge_key(sheet, row=row_index)
            if current_key != previous_key:
                self.merge_ozon_group(sheet, start_row=group_start, end_row=row_index - 1)
                group_start = row_index
                previous_key = current_key

        self.merge_ozon_group(sheet, start_row=group_start, end_row=sheet.max_row)

    def merge_ozon_group(self, sheet, *, start_row: int, end_row: int) -> None:
        """合并单个 Ozon 商品分组的重复列。"""

        if end_row <= start_row:
            return

        for column_index in self.EXCEL_OZON_MERGE_COLUMNS:
            sheet.merge_cells(
                start_row=start_row,
                start_column=column_index,
                end_row=end_row,
                end_column=column_index,
            )
            cell = sheet.cell(start_row, column_index)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    @classmethod
    def build_ozon_merge_key(cls, sheet, *, row: int) -> tuple[Any, ...]:
        """构造用于识别同一 Ozon 商品分组的键。"""

        return tuple(sheet.cell(row, column_index).value for column_index in cls.EXCEL_OZON_MERGE_COLUMNS)

    def write_report(
        self,
        *,
        source_reference: str,
        auth_state_path: Path,
        results: list[dict[str, Any]],
        excel_path: Path,
    ) -> Path:
        """把 1688 搜图结果写成 JSON 报告。"""

        output_dir = self.settings.ozon_scrape_output_path
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"alibaba1688_image_search_{time.strftime('%Y%m%d_%H%M%S')}.json"
        payload = {
            "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            "source_reference": source_reference,
            "auth_state_path": str(auth_state_path),
            "excel_path": str(excel_path),
            "results": results,
        }
        output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return output_path

    def save_to_database(self, payloads: list[dict[str, Any]]) -> dict[str, Any]:
        """把 1688 搜图结果保存到数据库。"""

        try:
            return self.repository.save_many(payloads)
        except Exception as exc:
            return {
                "status": "failed",
                "count": 0,
                "table": self.repository.table_name,
                "error": str(exc),
            }

    def save_to_sqlite(self, payloads: list[dict[str, Any]], *, source_reference: str) -> dict[str, Any]:
        """把 1688 搜图结果保存到 SQLite。"""

        sqlite_payloads = [{**payload, "source_reference": source_reference} for payload in payloads]
        try:
            return self.sqlite_repository.save_many(sqlite_payloads)
        except Exception as exc:
            return {
                "status": "failed",
                "count": 0,
                "error": str(exc),
            }

    def save_to_excel(self, rows: list[dict[str, Any]]) -> dict[str, Any]:
        """把 1688 搜图结果导出到 Excel。"""

        if not rows:
            return {
                "status": "skipped",
                "count": 0,
                "reason": "empty_rows",
            }

        try:
            output_path = self.export_to_excel(rows)
        except Exception as exc:
            return {
                "status": "failed",
                "count": 0,
                "error": str(exc),
            }

        return {
            "status": "saved",
            "count": len(rows),
            "path": str(output_path),
        }

    def mark_processed_products(self, product_ids: list[int], *, source_type: str) -> dict[str, Any]:
        """把本次已完成 1688 搜图的 Ozon 商品标记为已处理。"""

        if source_type != "sqlite_reviewed":
            return {
                "status": "skipped",
                "updated_count": 0,
                "reason": "non_sqlite_source",
            }

        return self.batch_repository.mark_products_alibaba_processed(product_ids)

    def compare_result_images_with_ai(self, ozon_product: dict[str, Any], items: list[dict[str, Any]]) -> dict[str, Any]:
        """先根据主图做 GPT 初筛，减少详情页访问量。"""

        if not items:
            return {"status": "skipped", "reason": "empty_items", "evaluated_items": 0, "passed_items": 0, "failed_items": 0}

        if not self.comparison_service.is_configured:
            for item in items:
                item["ai_image_comparison"] = {"status": "skipped", "reason": "openai_not_configured"}
            return {
                "status": "skipped",
                "reason": "openai_not_configured",
                "evaluated_items": 0,
                "passed_items": 0,
                "failed_items": 0,
            }

        evaluated = 0
        passed = 0
        failed = 0
        for item in items:
            try:
                item["ai_image_comparison"] = self.comparison_service.compare_product_images(
                    ozon_product=ozon_product,
                    supplier_product=item,
                )
                if item["ai_image_comparison"].get("status") == "completed":
                    evaluated += 1
                    if self.is_image_prefilter_passed(item["ai_image_comparison"]):
                        passed += 1
            except Exception as exc:
                failed += 1
                item["ai_image_comparison"] = {"status": "failed", "error": str(exc)}

        return {
            "status": "completed" if failed == 0 else "partial_failed",
            "evaluated_items": evaluated,
            "passed_items": passed,
            "failed_items": failed,
        }

    def select_best_image_match_items(self, items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """每个 Ozon 商品只保留通过主图初筛且分数最高的一条 1688 结果。"""

        if not items:
            return []

        passed_items = [
            item
            for item in items
            if self.is_image_prefilter_passed(item.get("ai_image_comparison") or {})
        ]
        if not passed_items:
            return []

        best_item = max(
            passed_items,
            key=lambda item: (
                int(((item.get("ai_image_comparison") or {}).get("image_match_score") or -1)),
                1 if (item.get("ai_image_comparison") or {}).get("same_product") else 0,
            ),
        )
        return [best_item]

    def is_image_prefilter_passed(self, comparison: dict[str, Any]) -> bool:
        """判断是否通过主图初筛。"""

        if comparison.get("status") != "completed":
            return False
        if comparison.get("same_product") is not True:
            return False
        return int(comparison.get("image_match_score") or 0) >= self.settings.alibaba1688_image_compare_pass_score

    @staticmethod
    def column_letter(index: int) -> str:
        """把列号转换成 Excel 列字母。"""

        letters = ""
        current = index
        while current > 0:
            current, remainder = divmod(current - 1, 26)
            letters = chr(65 + remainder) + letters
        return letters

    @staticmethod
    def format_attributes(attributes: list[dict[str, Any]] | None) -> str:
        """把 1688 属性列表压平成便于查看的文本。"""

        if not attributes:
            return ""
        return " | ".join(
            f"{str(item.get('key') or '').strip()}: {str(item.get('value') or '').strip()}"
            for item in attributes
            if item.get("key") and item.get("value")
        )

    @staticmethod
    def format_string_list(values: list[str] | None) -> str:
        """把字符串列表压平成 Excel 友好的文本。"""

        if not values:
            return ""
        return " | ".join(str(value).strip() for value in values if str(value).strip())

    @staticmethod
    def format_bool_value(value: bool | None) -> str:
        """把布尔值转成中文可读文案。"""

        if value is True:
            return "是"
        if value is False:
            return "否"
        return ""
